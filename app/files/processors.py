"""File processing — PDF, Word, Excel, PowerPoint, ZIP, JSON, XML, CSV."""
import json, csv, io, os
from app.utils.logger import logger

async def read_pdf(path: str) -> dict:
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(path)
        text = "\n".join([p.extract_text() or "" for p in reader.pages])
        return {"type":"pdf","pages":len(reader.pages),"text":text[:10000]}
    except Exception as e: return {"error": str(e)}

async def read_docx(path: str) -> dict:
    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join([p.text for p in doc.paragraphs])
        return {"type":"docx","text":text[:10000]}
    except Exception as e: return {"error": str(e)}

async def read_excel(path: str) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]; sheets[name] = [[str(c.value or "") for c in row] for row in ws.iter_rows(max_row=50)]
        return {"type":"excel","sheets": sheets}
    except Exception as e: return {"error": str(e)}

async def read_json_file(path: str) -> dict:
    try: 
        with open(path) as f: return {"type":"json","data": json.load(f)}
    except Exception as e: return {"error": str(e)}

async def read_csv_file(path: str) -> dict:
    try:
        with open(path, newline="") as f:
            reader = csv.DictReader(f); rows = list(reader)
        return {"type":"csv","columns":reader.fieldnames,"rows":rows[:100],"total_rows":len(rows)}
    except Exception as e: return {"error": str(e)}

FILE_HANDLERS = {".pdf": read_pdf, ".docx": read_docx, ".xlsx": read_excel, ".json": read_json_file, ".csv": read_csv_file}

async def process_file(path: str) -> dict:
    ext = os.path.splitext(path)[1].lower()
    handler = FILE_HANDLERS.get(ext)
    if not handler: return {"error": f"Unsupported: {ext}"}
    return await handler(path)
